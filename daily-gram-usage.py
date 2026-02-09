import pandas as pd
import xml.etree.ElementTree as ET
import re
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import os

tree = ET.parse('raport.xml')
root = tree.getroot()

documents = []

for document in root.findall('.//DokumentBaseWydrukOfDokumentRwDto'):
    document_body = document.find('Dokument')

    document_date = document_body.find('DataWydania').text if document_body.find('DataWydania') is not None else None
    document_description = document_body.find('Opis').text if document_body.find('Opis') is not None else None

    clients_matches = re.findall(r'posiłek|posiłki\s*(\d+)\s*x', document_description)
    clients = sum([int(client) for client in clients_matches[0:2] if client.isdigit()])

    positions_list = document_body.find('PozycjeDokumentuList')
    positions = positions_list.findall('PozycjaDokumentuDto') if positions_list is not None else []

    parsed_positions = []

    for position in positions:
        name = position.find('Nazwa').text
        unit = position.find('Jm').text
        amount = position.find('Ilosc').text

        amount_per_client = round(float(amount) / clients, 4)

        if unit == 'kg':
            amount_per_client_final = amount_per_client * 1000
        else:
            amount_per_client_final = None

        parsed_position = {
            'Nazwa produktu': name,
            'Jednostka': unit,
            'Ilość': amount,
            'Waga 1 szt w gramach': 'n.d.' if unit == 'kg' else None,
            'Ilość na 1 dziecko': amount_per_client,
            'Ilość na 1 dziecko (g)': amount_per_client_final
        }
        parsed_positions.append(parsed_position)

    documents.append({
        'date': document_date,
        'clients': clients,
        'positions': parsed_positions
    })

df = pd.DataFrame(documents)
df_exploded = df.explode("positions").reset_index(drop=True)
positions_df = pd.json_normalize(df_exploded["positions"])
positions_df = positions_df.reset_index(drop=True)
positions_df["date"] = df_exploded["date"]
positions_df["clients"] = df_exploded["clients"]

wb = Workbook()
wb.remove(wb.active)

for doc in documents:
    sheet_name = str(re.sub(r'[:\\/?*\[\]]', '-', doc['date'].split('T')[0])[:31])
    ws = wb.create_sheet(title=sheet_name)

    ws['A1'] = 'Data'
    ws['B1'] = doc['date'].split('T')[0]

    ws['A2'] = 'Liczba dzieci'
    ws['B2'] = doc['clients']

    df_positions = pd.DataFrame(doc['positions'])

    for r_idx, row in enumerate(dataframe_to_rows(df_positions, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if r_idx == 4:
                cell.value = value
            else:
                header = ws.cell(row=4, column=c_idx).value
                row_unit = ws.cell(row=r_idx, column=[cell.column for cell in ws[4] if cell.value == 'Jednostka'][0]).value

                if header == 'Ilość na 1 dziecko (g)':
                    if row_unit == 'kg':
                        cell.value = row[c_idx - 1]
                    else:
                        col_ilosc_g = [cell.column for cell in ws[4] if cell.value == 'Waga 1 szt w gramach'][0]
                        col_na_1 = [cell.column for cell in ws[4] if cell.value == 'Ilość na 1 dziecko'][0]
                        cell.value = f"={chr(64 + col_ilosc_g)}{r_idx}*{chr(64 + col_na_1)}{r_idx}"
                else:
                    cell.value = value

filepath = "/mnt/c/Users/jbednarek/Desktop/daily_documents.xlsx"
if os.path.exists(filepath):
    os.remove(filepath)
wb.save(filepath)
